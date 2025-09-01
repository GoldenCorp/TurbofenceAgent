from fastapi import APIRouter, Depends, File, Form, Query, Request, UploadFile
from vos import (
    BaojiaInputModel,
    BaojiaStrnputModel
)
# import httpx
from fastapi.security import HTTPBearer
import openpyxl
from spire.xls import *
import time
import datetime
from qcloud_cos import CosConfig, CosS3Client
# Scheme for the Authorization header
token_auth_scheme = HTTPBearer()
# from http_exception import HTTPException
baojiaRouter = APIRouter(prefix='/baojia')

@baojiaRouter.post("/baojia", operation_id="baojia", summary="这个工具可以针对周界安防项目，接受计算的报价结果，形成报价方案文件，提供该文件的下载链接提供用户下载使用。")
def baojia(chuchaiParam:BaojiaStrnputModel,token=Depends(token_auth_scheme)):
    print("0")
    exTemplatePath = 'baojia3.xlsx'
    wb = openpyxl.load_workbook(exTemplatePath)

    dateNow = datetime.datetime.now()
    str_dateNow = dateNow.strftime("%Y%m%d")
    orderNum = dateNow.strftime("%Y%m%d%H%M%S")
    sheet = wb.active
    sheet['F11'] = chuchaiParam.totalCamera
    sheet['F12'] = chuchaiParam.totalCameraPole
    sheet['F21'] = chuchaiParam.totalLength
    sheet['G3'] = str_dateNow
    sheet['C3'] = orderNum
    
    timestamp = int(time.time())
    xlsfilename = f'baojia_{timestamp}.xlsx'
    wb.save(xlsfilename)
    print("wb.save(xlsfilename)")
    try:
        workbook = Workbook()
        print("workbook = Workbook()")
        workbook.LoadFromFile(xlsfilename)
        pdffilename = f"baojia_{timestamp}.pdf"
        print(pdffilename)
        workbook.SaveToFile(pdffilename, FileFormat.PDF)
        workbook.Dispose()
    except Exception as e:
        print(f"Error occurred: {e}")
        return {"error": str(e)}
    
    secret_id = 'AKIDdAUfgV6UebLrAG9tQltnX5UilcGhy2tX'  # 替换为实际SecretId
    secret_key = 'hO0TjgYDfvfHgsrfGepoMoHhHg6NWSBh'     # 替换为实际SecretKey
    region = 'ap-beijing'     # 存储桶地域
    bucket = 'baojia-1370877121'  # 存储桶名称

    config = CosConfig(Region=region, SecretId=secret_id, SecretKey=secret_key)
    client = CosS3Client(config)
    download_url=""
    #上传文件
    response = client.upload_file(
        Bucket=bucket,
        Key=pdffilename,  # 对象键（上传后的文件名）
        LocalFilePath=pdffilename  # 本地文件路径
    )
    download_url = client.get_presigned_url(
        Method='GET',
        Bucket=bucket,
        Key=pdffilename,
        Expired=int(time.time()) + 600  # 10分钟后过期
    )
    return download_url